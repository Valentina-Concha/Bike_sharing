#NUEVOO CODIGO QUE AGREGA LA COLUMNA
#pandas es una biblioteca Python para manipular y analizar datos sobre todo df, especialmente archivos como Excel.
import pandas as pd
#openpyxl es una biblioteca que permite trabajar directamente con archivos Excel. Aquí se usa para cargar y actualizar el archivo existente.
from openpyxl import load_workbook
#para importar los datos de bicicletas de py
#los datos que entrega en los links de la web son tipo json
import pybikes
#para exportar la fecha actual de hoy
from datetime import date
from datetime import datetime
import pytz
import requests
import pandas as pd
#import requests: Se importa la librería requests, que se utiliza para realizar peticiones HTTP en Python. Esto permite acceder a datos de APIs (interfaces de programación de aplicaciones) que suelen ofrecer información en formato JSON.
import requests
import pandas as pd
from datetime import date, datetime
from scipy.stats import skew, kurtosis

timezone=pytz.timezone('America/Santiago')

#station_info_url: Esta variable almacena la URL que proporciona información sobre las estaciones de bicicletas, como nombre, capacidad y ubicación.
#station_status_url: Esta variable almacena la URL que proporciona el estado actual de las estaciones, incluyendo la disponibilidad de bicicletas y espacios.
# URL de la API de información de estaciones
station_info_url = 'https://santiago.publicbikesystem.net/customer/ube/gbfs/v1/en/station_information'

# URL de la API de estado de estaciones
station_status_url = 'https://santiago.publicbikesystem.net/customer/ube/gbfs/v1/en/station_status'

# Hacer las solicitudes GET
#Se realizan dos solicitudes HTTP GET a las URLs definidas anteriormente. Estas solicitudes recuperan datos de las APIs.
station_info_response = requests.get(station_info_url)
station_status_response = requests.get(station_status_url)


# Verificar si ambas solicitudes fueron exitosas. Aquí se comprueba si ambas solicitudes fueron exitosas (código de estado 200 significa que la solicitud fue exitosa)
if station_info_response.status_code == 200 and station_status_response.status_code == 200:
    #Cargar los datos de las respuestas en formato JSON. Si las solicitudes fueron exitosas, los datos se cargan en formato JSON a las variables correspondientes (station_info_data y station_status_data)
    station_info_data = station_info_response.json()
    station_status_data = station_status_response.json()

    # Listas para almacenar los datos. Se inicializan listas vacías para almacenar los datos de cada estación que se extraerán del JSON.
    estaciones = []
    grupos = []
    capacidad = []
    latitud = []
    longitud = []
    bikes = []
    free = []
    num_bikes_disabled = []
    fechas = []  # Lista para almacenar la fecha
    horas = []   # Lista para almacenar la hora

    # Obtener la fecha y la hora actuales
    fecha_actual = date.today(timezone)
    hora_actual = datetime.now(timezone)

    # Obtener información de estaciones
    #stations_info: Extrae la información sobre las estaciones desde los datos JSON.
    #Ciclo for: Itera sobre cada estación en stations_info para extraer los detalles.
    #station.get('station_id', 'Unknown ID'): Intenta obtener el station_id. Si no está disponible, se asigna el valor 'Unknown ID'. Por esto puse "Unknown", es una forma de manejar posibles valores faltantes o no disponibles.
    #Se obtienen otras propiedades de la estación (nombre, grupos, capacidad, latitud y longitud) de manera similar, utilizando valores por defecto en caso de que falten.
    stations_info = station_info_data.get('data', {}).get('stations', [])

    for station in stations_info:
        station_id = station.get('station_id', 'Unknown ID')
        estaciones.append(station.get('name', 'Unknown Station'))
        grupos.append(station.get('groups', ['Sin grupo'])[0])  # Si no hay grupo, colocar 'Sin grupo'
        capacidad.append(station.get('capacity', 'Unknown Capacity'))
        latitud.append(station.get('lat', 'Unknown Latitude'))
        longitud.append(station.get('lon', 'Unknown Longitude'))
        fechas.append(fecha_actual)  # Agregar la fecha a cada fila
        horas.append(hora_actual)    # Agregar la hora a cada fila

    # Obtener estado de estaciones
    #Se repite un proceso similar para el estado de las estaciones:
    #stations_status: Extrae los datos de estado desde el JSON.
    #status_dict: Se crea un diccionario que relaciona cada station_id con su estado.
    #Luego se extraen datos sobre el número de bicicletas disponibles, espacios libres y bicicletas deshabilitadas, utilizando el mismo enfoque de valores por defecto.
    stations_status = station_status_data.get('data', {}).get('stations', [])

    #Crear un diccionario para relacionar 'station_id' con los datos de estado
    #Se repite un proceso similar para el estado de las estaciones:
    #stations_status: Extrae los datos de estado desde el JSON.
    #status_dict: Se crea un diccionario que relaciona cada station_id con su estado.
    #Luego se extraen datos sobre el número de bicicletas disponibles, espacios libres y bicicletas deshabilitadas, utilizando el mismo enfoque de valores por defecto.
    status_dict = {station.get('station_id'): station for station in stations_status}

    for station in stations_info:
        station_id = station.get('station_id', 'Unknown ID')
        status = status_dict.get(station_id, {})
        bikes.append(status.get('num_bikes_available', 'Unknown Bikes'))
        free.append(status.get('num_docks_available', 'Unknown Free'))
        num_bikes_disabled.append(status.get('num_bikes_disabled', 'Unknown Disabled'))

    # Crear el DataFrame final con todas las columnas necesarias
    df = pd.DataFrame({
        'Estacion': estaciones,
        'Comuna': grupos,
        'Capacidad (Slots)': capacidad,
        'Latitud': latitud,
        'Longitud': longitud,
        'Bikes disponibles': bikes,
        'Free': free,
        'Bikes Deshabilitadas': num_bikes_disabled,
        'Fecha': fechas,  # Agregar la columna Fecha
        'Hora': horas     # Agregar la columna Hora
    })


    # Mostrar en Python el número total de estaciones descargadas
    total_estaciones = len(stations_info)
    print(f"Número total de estaciones descargadas: {total_estaciones}")

    # Si quieres asegurarte de que se descargaron ambas listas de datos correctamente
    print(f"Número de estaciones en la información de estaciones: {len(stations_info)}")
    print(f"Número de estaciones en el estado de estaciones: {len(stations_status)}")

else:
    print(f"Error en las solicitudes: {station_info_response.status_code}, {station_status_response.status_code}")
print(df)

#GUARDA LOS DATOS EN BD DENTRO DEL MISMO EXCEL

# Definir el nombre del archivo
#Para que lo descargue vinculando la misma carpeta del Git hub que es bike_sharing
file_name = 'bd_estaciones.xlsx'

# Crear el DataFrame (puedes usar tu propio DataFrame aquí)
# df_data_countries = pd.DataFrame({"Columna1": [5, 6], "Columna2": [7, 8]})

# Definir el nombre de la hoja donde vas a escribir
sheet_name = 'BD2'

# Cargar el archivo de Excel existente
book = load_workbook(file_name)
if sheet_name in book.sheetnames:
    # Cargar los datos existentes de la hoja especificada en un DataFrame
    existing_df = pd.read_excel(file_name, sheet_name=sheet_name)

    # Eliminar cualquier columna completamente vacía, para que no se agreguen columnas como Unnamed: 1, 2..
    existing_df = existing_df.dropna(axis=1, how='all')

    # Agregar los nuevos datos al final del DataFrame existente
    updated_df = pd.concat([existing_df, df], ignore_index=True)

    # Guardar el DataFrame actualizado en la misma hoja, sobrescribiéndola
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
else:
    # Si la hoja no existe, escribir el DataFrame desde cero
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Datos exportados a {file_name} en la hoja '{sheet_name}'")


